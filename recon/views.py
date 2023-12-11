import io
import json
import logging
import os
import datetime as dt
from zipfile import ZipFile

from django.http import FileResponse, HttpResponse, Http404
from django.views import View
from django.db.models import Q, F, Case, When, Value, CharField
from django.db.models.functions import Cast
from django.http import HttpResponse

import pandas as pd
from openpyxl import load_workbook
from rest_framework import generics, status, viewsets
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from recon.index import reconcileMain
from recon.setlement_ import setleSabs, settle
from recon.utils import unserializable_floats
from .models import Recon, ReconLog, UploadedFile, Bank, UserBankMapping, Transactions
from .serializers import (
    ReconcileSerializer, ReconciliationSerializer, SabsSerializer,
    SettlementSerializer, UploadedFileSerializer, LogSerializer, TransactionSerializer
)

current_date = dt.date.today().strftime('%Y-%m-%d')
# Get the current date and time
current_datetime = dt.datetime.now()
current_day = current_datetime.replace(hour=23, minute=59, second=59, microsecond=999999)

class CustomReconciliationError(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)

class CustomFileIOError(Exception):
    def __init__(self, message):
        self.message = message
        super().__init__(self.message)
# Create your views here.

def get_swift_code_from_request(request):
    user = request.user
    mapping = UserBankMapping.objects.filter(user=user)[0]
    bank = mapping.bank
    swift_code = bank.swift_code

    return swift_code

def get_bank_code_from_request(request):
    user = request.user
    mapping = UserBankMapping.objects.filter(user=user)[0]
    bank = mapping.bank
    bank_code = bank.bank_code    
    
    return bank_code

def get_username_from_request(request):
    user = request.user
    username = user.username
    return username

class UploadedFilesViewset(viewsets.ModelViewSet):
    queryset = UploadedFile.objects.all()
    serializer_class = UploadedFileSerializer
    def create(self, request, *args, **kwargs):
        user = request.user
        file = request.FILES['file']
        wb = load_workbook(file)
        sheet = wb.get_sheet_by_name("Sheet1")
        start_row = 2
        count = 0
        for _ in sheet.iter_rows(min_row=start_row,max_row=10):
            row_no = start_row+count
            time = sheet.cell(row_no,1).value
            transaction_type = sheet.cell(row_no,2).value
            amount = sheet.cell(row_no,3).value
            abc_reference = sheet.cell(row_no,4).value
            recon = Recon(
                date_time=time,
                last_modified_by_user=user,
                trn_ref=abc_reference,
                )
            recon.save()
            print(time,transaction_type,amount,abc_reference)
            count+=1
        
        return super().create(request, *args, **kwargs) 

class ReconcileView(APIView):
    serializer_class = ReconcileSerializer
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        user = request.user
        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            bank_code = get_bank_code_from_request(request)

            # Save the uploaded file temporarily
            temp_file_path = "temp_file.xlsx"
            try:
                with open(temp_file_path, "wb") as buffer:
                    buffer.write(uploaded_file.read())

                try:
                    # Call the main function with the path of the saved file and the swift code
                    merged_df, reconciled_data, succunreconciled_data, exceptions, feedback, requestedRows, UploadedRows, date_range_str = reconcileMain(
                        temp_file_path, bank_code, user)

                    # Perform clean up: remove the temporary file after processing
                    os.remove(temp_file_path)

                    data = {
                        "reconciledRows": len(reconciled_data) if reconciled_data is not None else 0,
                        "unreconciledRows": len(succunreconciled_data) if succunreconciled_data is not None else 0,
                        "exceptionsRows": len(exceptions) if exceptions is not None else 0, 
                        "feedback": feedback,
                        "RequestedRows": requestedRows,
                        "UploadedRows": UploadedRows,
                        "min_max_DateRange": date_range_str,
                        "reconciled_data": reconciled_data.to_dict(orient='records') if isinstance(reconciled_data, pd.DataFrame) else reconciled_data,
                        "succunreconciled_data": succunreconciled_data.to_dict(orient='records') if isinstance(succunreconciled_data, pd.DataFrame) else succunreconciled_data
                    }
                    return Response(data, status=status.HTTP_200_OK)

                except Exception as e:
                    # If there's an error during the process, ensure the temp file is removed
                    if os.path.exists(temp_file_path):
                        os.remove(temp_file_path)

                    # Handle the specific exceptions and raise custom exceptions with additional context
                    logging.error(f"An error occurred during reconciliation: {str(e)}")
                    raise CustomReconciliationError(f"Error during reconciliation: {str(e)}")

            except IOError as ioe:
                # Handle file I/O errors
                logging.error(f"An error occurred while saving the uploaded file: {str(ioe)}")
                raise CustomFileIOError(f"Error saving the uploaded file: {str(ioe)}")

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ReversalsView(generics.ListAPIView):
    serializer_class = TransactionSerializer

    def get_queryset(self):
        bank_code = get_bank_code_from_request(self.request)        

        queryset = Transactions.objects.filter(
            Q(request_type__in=['1420', '1421']) &
            ~Q(txn_type__in=['BI', 'MINI']) & 
            ~Q(processing_code__in=['320000', '340000', '510000', '370000', '180000','360000']) & ~Q(amount='0') & 
            (Q(issuer_code=bank_code) | Q(acquirer_code=bank_code)) & ~Q(response_code='00') & Q(date_time=current_day) 
        ).annotate(
            Reversal_type=Case(
                When(request_type='1420', then=Value('Reversal')),
                When(request_type='1421', then=Value('Repeat Reversal')),
                default=Value(None),
                output_field=CharField()
            ),
            Status=Case(
                When(response_code=None, then=Value('Pending')),
                When(response_code='00', then=Value('Successful')),
                default=Value('Failed'),
                output_field=CharField()
            )
        ).values(
            'date_time',
            'txn_id',
            'trn_ref',
            'amount',
            'issuer',
            'acquirer',
            'txn_type',
            'Reversal_type',
            'Status'
        ).distinct()
        
        return queryset

class ExceptionsView(generics.ListAPIView):
       
    serializer_class = ReconciliationSerializer
    """
    Retrieve Exceptions data.
    """

    def get_queryset(self):
        # Use values from .env for database connection
        bank_code = get_bank_code_from_request(self.request)
        return Recon.objects.filter(Q(excep_flag="Y")& (Q(issuer_code = bank_code)|Q(acquirer_code = bank_code)))

class ReconStatsView(generics.ListAPIView):
    serializer_class = LogSerializer
    """
    Retrieve Stats data.
    """

    def get_queryset(self):
        # Use values from .env for database connection
        bank_code = get_bank_code_from_request(self.request)
        return ReconLog.objects.filter(Q(bank_id=bank_code))  
        
class sabsreconcile_csv_filesView(APIView):

    serializer_class = SabsSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            uploaded_file = serializer.validated_data['file']
            batch_number = serializer.validated_data['batch_number']

            # Save the uploaded file temporarily
            temp_file_path = "temp_file.xlsx"
            with open(temp_file_path, "wb") as buffer:
                buffer.write(uploaded_file.read())

            try:
                # Assume setleSabs returns dataframes as one of its outputs
                _, matched_setle, _, unmatched_setlesabs = setleSabs(temp_file_path, batch_number)
                
                # Perform clean up: remove the temporary file after processing
                os.remove(temp_file_path)

                matched_csv = matched_setle.to_csv(index=False)
                unmatched_csv = unmatched_setlesabs.to_csv(index=False)

                # Create a zip file in memory
                memory_file = io.BytesIO()
                with ZipFile(memory_file, 'w') as zf:
                    zf.writestr('matched_setle.csv', matched_csv)
                    zf.writestr('unmatched_setlesabs.csv', unmatched_csv)
                
                # u will figure ou how to retun a zipped file here
                memory_file.seek(0)

                response = FileResponse(memory_file, content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=Settlement_.zip'
                return response            
            
            except Exception as e:
                # If there's an error during the process, ensure the temp file is removed
                if os.path.exists(temp_file_path):
                    os.remove(temp_file_path)
                
                # Return error as response
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class SettlementView(APIView):
    serializer_class = SettlementSerializer

    def post(self, request):
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            batch_number = serializer.validated_data['batch_number']

            try:
                # Assume the settle function is defined and available here
                settlement_result = settle(batch_number)

                # Handle case where no records were found or an error occurred in settle
                if settlement_result is None or settlement_result.empty:
                    return Response({"detail": "No records for processing found or an error occurred."},
                                    status=status.HTTP_400_BAD_REQUEST)

                # Convert the DataFrame to CSV
                settlement_csv = settlement_result.to_csv(index=False)

                # Create a zip file in memory
                memory_file = io.BytesIO()
                with ZipFile(memory_file, 'w') as zf:
                    zf.writestr('settlement_result.csv', settlement_csv)

                memory_file.seek(0)

                response = FileResponse(memory_file, content_type='application/zip')
                response['Content-Disposition'] = 'attachment; filename=Settlement_.zip'
                return response

            except Exception as e:
                # Handle other unexpected errors
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

